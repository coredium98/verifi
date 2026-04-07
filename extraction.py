import io
import math
import pdfplumber
import pandas as pd
from openpyxl import load_workbook

def _normalize(text):
    if text is None:
        return ""
    return str(text).lower().strip().lstrip()

def _parse_number(val):
    if val is None:
        return None
    if isinstance(val, (int, float)):
        if math.isnan(val):
            return None
        return float(val)
    s = str(val).replace(",","").replace("$","").replace("(", "-").replace(")","").replace("%","").strip()
    try:
        return float(s)
    except ValueError:
        return None

# Ordered rules — first match wins. More specific patterns go first.
MATCH_RULES = [
    ("operating_cash_flow", ["operating cash flow","cash from operat","cash flow from operat"], []),
    ("ebitda",              ["ebitda"],                                                          []),
    ("gross_profit",        ["gross profit","gross margin"],                                     ["margin %", "margin%"]),
    ("revenue",             ["revenue / sales","net sales","total revenue","revenue"],           ["cost","other revenue"]),
    ("cogs",                ["cost of goods","cogs","cost of sales","direct cost"],              []),
    ("net_income",          ["net income","net profit","net earnings","net loss"],               ["change in","add:"]),
    ("owner_comp",          ["owner / officer","owner/officer","officer compensation",
                             "officer salary","owner compensation","owner salary"],              []),
    ("accounts_receivable", ["accounts receivable","trade receivable"],                          ["change in","change -"]),
    ("depreciation",        ["depreciation","amortization"],                                    ["accumulated","add:"]),
    ("interest",            ["interest expense"],                                               []),
    ("tax",                 ["income tax","tax expense"],                                       []),
]

def _classify(label):
    t = _normalize(label)
    if not t:
        return None
    for key, required, excluded in MATCH_RULES:
        if any(r in t for r in required):
            if not any(e in t for e in excluded):
                return key
    return None

def _better(existing, new):
    """Return whichever list has more non-zero, non-None values."""
    def score(lst):
        return sum(1 for v in lst if v is not None and v != 0)
    return new if score(new) > score(existing) else existing

def extract_from_excel(file_bytes):
    data = {}
    try:
        wb = load_workbook(file_bytes, data_only=True)
        ws = wb.active
    except Exception:
        try:
            df = pd.read_excel(file_bytes, header=None)
            return _from_df(df)
        except Exception:
            return data

    for row in ws.iter_rows():
        label = row[0].value if row else None
        if label is None:
            continue
        key = _classify(str(label))
        if key is None:
            continue
        nums = []
        for cell in row[1:]:
            v = cell.value
            # skip error strings and formula strings
            if isinstance(v, str) and (v.startswith("#") or v.startswith("=")):
                continue
            n = _parse_number(v)
            if n is not None:
                nums.append(n)
        nums = nums[:3]
        if not nums:
            continue
        if key in data:
            data[key] = _better(data[key], nums)
        else:
            data[key] = nums

    return data

def _from_df(df):
    data = {}
    for _, row in df.iterrows():
        label = row.iloc[0]
        if label is None:
            continue
        key = _classify(str(label))
        if key is None:
            continue
        nums = [_parse_number(v) for v in row.iloc[1:] if _parse_number(v) is not None][:3]
        if nums:
            if key in data:
                data[key] = _better(data[key], nums)
            else:
                data[key] = nums
    return data

def extract_from_pdf(file_bytes):
    data = {}
    with pdfplumber.open(file_bytes) as pdf:
        for page in pdf.pages:
            for table in (page.extract_tables() or []):
                for row in table:
                    if not row or row[0] is None:
                        continue
                    key = _classify(str(row[0]))
                    if key is None:
                        continue
                    nums = [_parse_number(c) for c in row[1:] if _parse_number(c) is not None][:3]
                    if nums:
                        if key in data:
                            data[key] = _better(data[key], nums)
                        else:
                            data[key] = nums
    return data

def assess_quality(data, years_claimed):
    has_revenue   = "revenue" in data
    has_gp        = "gross_profit" in data or ("cogs" in data and has_revenue)
    has_ni        = "net_income" in data
    has_income    = has_revenue and has_gp and has_ni
    has_balance   = "accounts_receivable" in data
    has_cash_flow = "operating_cash_flow" in data

    core_count = sum([has_revenue, has_gp, has_ni, has_balance, has_cash_flow])

    if core_count >= 4:
        score = "Sufficient"
    elif core_count >= 2:
        score = "Partial"
    else:
        score = "Insufficient"

    missing = []
    if not has_income:
        missing.append("complete income statement (revenue, gross profit, net income)")
    if not has_balance:
        missing.append("balance sheet with accounts receivable")
    if not has_cash_flow:
        missing.append("cash flow statement or operating cash flow figure")
    if "owner_comp" not in data:
        missing.append("owner / officer compensation detail")

    return {
        "has_income_stmt": has_income,
        "has_balance":     has_balance,
        "has_cash_flow":   has_cash_flow,
        "score":           score,
        "missing":         missing,
        "found_keys":      list(data.keys()),
    }

MANUAL_FIELDS = [
    ("revenue",             "Revenue"),
    ("cogs",                "Cost of Goods Sold"),
    ("gross_profit",        "Gross Profit"),
    ("net_income",          "Net Income"),
    ("operating_cash_flow", "Operating Cash Flow"),
    ("accounts_receivable", "Accounts Receivable (Balance Sheet)"),
    ("owner_comp",          "Owner / Officer Compensation"),
    ("ebitda",              "EBITDA"),
]
